import urllib
import os
import csv
from openpyxl import *
quotes = ["601988.SS","600028.SS"]
for x in range(len(quotes)):
    quote = quotes[x]
    urllib.urlretrieve ("http://real-chart.finance.yahoo.com/table.csv?s=%s&d=3&e=10&f=2016&g=d&a=6&b=5&c=2014&ignore=.csv" % str(quote),"%s.csv" % str(quote))
    with open("%s.csv"%quote,'rb') as csvfile:

        reader = csv.DictReader(csvfile)
        for row in reader:
            print row['Open'] +" " +row['Date']
os.remove('%s.csv' % quote)

"""
wb =load_workbook('/Users/elibaum/Desktop/Testshares.xlsx')
ws = wb.active
with open("%s.csv"%quote,'rb') as csvfile:
    reader = csv.DictReader(csvfile)
    for row in reader:
        print row['Open'] +" " +row['Date']
        for x in range(1,23):
            tested = 'C' +str(x)
            if str(ws[tested].value) == 'None':
                ws[tested] = row['Open']
                tempdate= row['Date']
                ws['D'+str(x)] =tempdate
                break
wb.save('/Users/elibaum/Desktop/Testshares.xlsx')
os.remove('%s.csv' % quote)
"""